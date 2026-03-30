"""Export matched results to a user-friendly Excel report."""

import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from crawlers.reporter import _format_budget
from utils.config_loader import load_config


def load_latest_matched(config_path: str | None = None) -> dict:
    """Find and load the latest ace_matched_*.json file via run tracker."""
    from utils.run_tracker import get_latest_file

    cfg = load_config(config_path)
    data_dir = cfg["output"]["data_dir"]

    matched_file = get_latest_file(data_dir, "matched")
    if not matched_file:
        print("找不到匹配資料，請先執行 python main.py --all")
        sys.exit(1)
    with open(matched_file, "r", encoding="utf-8") as f:
        return json.load(f), Path(matched_file).name


def create_excel_report(data: dict, source_file: str, config_path: str | None = None):
    """Create a formatted Excel report."""
    wb = Workbook()

    # ── Colors & Styles ──
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(name="Microsoft JhengHei", bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name="Microsoft JhengHei", bold=True, size=14, color="1F4E79")
    SUBTITLE_FONT = Font(name="Microsoft JhengHei", size=10, color="666666")
    BODY_FONT = Font(name="Microsoft JhengHei", size=10)
    LINK_FONT = Font(name="Microsoft JhengHei", size=10, color="0563C1", underline="single")
    HIGH_FILL = PatternFill("solid", fgColor="E2EFDA")
    MEDIUM_FILL = PatternFill("solid", fgColor="FFF2CC")
    LOW_FILL = PatternFill("solid", fgColor="F2F2F2")
    BU1_FONT = Font(name="Microsoft JhengHei", size=10, bold=True, color="2E75B6")
    BU2_FONT = Font(name="Microsoft JhengHei", size=10, bold=True, color="C55A11")
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="D9D9D9"),
    )
    WRAP = Alignment(wrap_text=True, vertical="center")
    CENTER = Alignment(horizontal="center", vertical="center")

    tenders = data.get("tenders", [])
    passed = [t for t in tenders if t.get("passes_filter")]
    passed.sort(key=lambda t: t.get("fit_score", 0), reverse=True)

    # ═══════════════════════════════════════════
    # Sheet 1: 商機總覽 (Overview)
    # ═══════════════════════════════════════════
    ws = wb.active
    ws.title = "商機總覽"
    ws.sheet_properties.tabColor = "1F4E79"

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "ACE-SPIDER 採購商機分析報告"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    crawl_time = data.get("crawl_time", "")[:19].replace("T", " ")
    ws.merge_cells("A2:L2")
    ws["A2"] = f"報告時間：{datetime.now().strftime('%Y-%m-%d %H:%M')}　｜　爬取時間：{crawl_time}　｜　來源：{source_file}"
    ws["A2"].font = SUBTITLE_FONT
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        ("排名", 6),
        ("優先級", 8),
        ("BU", 8),
        ("標案名稱", 45),
        ("機關名稱", 22),
        ("預算金額", 14),
        ("Fit 分數", 10),
        ("ROI", 8),
        ("技術標籤", 25),
        ("產業標籤", 18),
        ("AI 摘要", 40),
        ("匹配聯繫人", 30),
        ("截止日期", 14),
        ("連結", 12),
    ]

    row = 4
    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 24

    # Freeze panes
    ws.freeze_panes = "A5"
    # Auto filter
    ws.auto_filter.ref = f"A4:N{4 + len(passed)}"

    # Data rows
    for i, t in enumerate(passed, 1):
        row = 4 + i
        tag = t.get("tag_result") or {}
        contacts = t.get("matched_contacts", [])
        budget = t.get("budget")
        priority = t.get("priority", "low")
        bu = tag.get("bu_assignment", "none")

        # Priority display
        pri_map = {"high": "🔴 高", "medium": "🟡 中", "low": "⚪ 低"}
        pri_text = pri_map.get(priority, priority)

        budget_str = _format_budget(budget) if budget else "未公開"

        # BU display
        bu_map = {"BU1": "BU1 區塊鏈", "BU2": "BU2 AI", "both": "BU1+BU2", "none": "-"}
        bu_text = bu_map.get(bu, bu)

        # Contacts
        contact_strs = []
        for c in contacts[:3]:
            contact_strs.append(f"{c['name']}（{c.get('company', '')}，{c.get('position', '')}）")
        contact_text = "\n".join(contact_strs)

        # Tech/Industry tags
        tech_text = "、".join(tag.get("tech_tags", []))
        industry_text = "、".join(tag.get("industry_tags", []))

        # URL
        url = t.get("url", "")

        values = [
            i,
            pri_text,
            bu_text,
            t.get("tender_name", ""),
            t.get("org_name", ""),
            budget_str,
            t.get("fit_score", 0),
            t.get("roi_score", 0),
            tech_text,
            industry_text,
            tag.get("relevance_summary", ""),
            contact_text,
            t.get("deadline", ""),
            "點擊查看",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP if col in (4, 9, 10, 11, 12) else CENTER

        # Style: priority color
        row_fill = {"high": HIGH_FILL, "medium": MEDIUM_FILL}.get(priority, LOW_FILL)
        for col in range(1, len(values) + 1):
            ws.cell(row=row, column=col).fill = row_fill

        # Style: BU color
        bu_cell = ws.cell(row=row, column=3)
        if bu in ("BU1", "both"):
            bu_cell.font = BU1_FONT
        elif bu == "BU2":
            bu_cell.font = BU2_FONT

        # Hyperlink
        if url:
            link_cell = ws.cell(row=row, column=14)
            link_cell.hyperlink = url
            link_cell.font = LINK_FONT

        ws.row_dimensions[row].height = 50

    # ═══════════════════════════════════════════
    # Sheet 2: 統計摘要 (Summary)
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("統計摘要")
    ws2.sheet_properties.tabColor = "2E75B6"

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 30

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "統計摘要"
    ws2["A1"].font = TITLE_FONT
    ws2.row_dimensions[1].height = 30

    # Overall stats
    stats = [
        ("總爬取標案數", len(tenders)),
        ("通過篩選數", len(passed)),
        ("篩選通過率", f"{len(passed)/max(len(tenders),1)*100:.1f}%"),
        ("", ""),
        ("BU1（區塊鏈金流追蹤）", sum(1 for t in passed if (t.get("tag_result") or {}).get("bu_assignment") in ("BU1", "both"))),
        ("BU2（AI 應用客製）", sum(1 for t in passed if (t.get("tag_result") or {}).get("bu_assignment") in ("BU2", "both"))),
        ("無關聯", sum(1 for t in passed if (t.get("tag_result") or {}).get("bu_assignment") == "none")),
        ("", ""),
        ("🔴 高優先級", sum(1 for t in passed if t.get("priority") == "high")),
        ("🟡 中優先級", sum(1 for t in passed if t.get("priority") == "medium")),
        ("⚪ 低優先級", sum(1 for t in passed if t.get("priority") == "low")),
        ("", ""),
        ("平均 Fit 分數", f"{sum(t.get('fit_score',0) for t in passed)/max(len(passed),1):.1f}"),
        ("總預算金額", f"{sum(t.get('budget',0) or 0 for t in passed)/100_000_000:.1f} 億"),
        ("GPT Token 消耗", f"{data.get('total_tokens_used', 0):,}"),
    ]

    for i, (label, value) in enumerate(stats, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Microsoft JhengHei", bold=True, size=11)
        ws2.cell(row=i, column=2, value=value).font = Font(name="Microsoft JhengHei", size=11)

    # Top 5 by ROI
    ws2.merge_cells("A20:D20")
    ws2["A20"] = "TOP 5 最高 ROI 標案"
    ws2["A20"].font = Font(name="Microsoft JhengHei", bold=True, size=12, color="1F4E79")

    top_headers = ["標案名稱", "預算", "ROI", "BU"]
    for col, h in enumerate(top_headers, 1):
        cell = ws2.cell(row=21, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    top_roi = sorted(passed, key=lambda t: t.get("roi_score", 0), reverse=True)[:5]
    for i, t in enumerate(top_roi, 22):
        tag = t.get("tag_result") or {}
        budget = t.get("budget", 0) or 0
        ws2.cell(row=i, column=1, value=t.get("tender_name", "")).font = BODY_FONT
        ws2.cell(row=i, column=2, value=_format_budget(budget)).font = BODY_FONT
        ws2.cell(row=i, column=3, value=t.get("roi_score", 0)).font = BODY_FONT
        ws2.cell(row=i, column=4, value=tag.get("bu_assignment", "")).font = BODY_FONT

    # ═══════════════════════════════════════════
    # Sheet 3: BU1 專屬 & Sheet 4: BU2 專屬
    # ═══════════════════════════════════════════
    for bu_id, bu_name, tab_color in [("BU1", "BU1 區塊鏈", "C55A11"), ("BU2", "BU2 AI應用", "2E75B6")]:
        ws_bu = wb.create_sheet(bu_name)
        ws_bu.sheet_properties.tabColor = tab_color

        bu_tenders = [t for t in passed if (t.get("tag_result") or {}).get("bu_assignment") in (bu_id, "both")]
        bu_tenders.sort(key=lambda t: t.get("fit_score", 0), reverse=True)

        ws_bu.merge_cells("A1:H1")
        ws_bu["A1"] = f"{bu_name} 相關商機（{len(bu_tenders)} 筆）"
        ws_bu["A1"].font = TITLE_FONT
        ws_bu.row_dimensions[1].height = 30

        bu_headers = [
            ("排名", 6), ("標案名稱", 45), ("機關名稱", 22), ("預算", 14),
            ("Fit", 8), ("ROI", 8), ("技術標籤", 30), ("匹配聯繫人", 30), ("連結", 12),
        ]

        for col, (name, width) in enumerate(bu_headers, 1):
            cell = ws_bu.cell(row=3, column=col, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            ws_bu.column_dimensions[get_column_letter(col)].width = width

        ws_bu.freeze_panes = "A4"

        for i, t in enumerate(bu_tenders, 1):
            row = 3 + i
            tag = t.get("tag_result") or {}
            contacts = t.get("matched_contacts", [])
            budget = t.get("budget", 0) or 0
            url = t.get("url", "")

            contact_text = "、".join(f"{c['name']}（{c.get('position', '')}）" for c in contacts[:2])

            ws_bu.cell(row=row, column=1, value=i).font = BODY_FONT
            ws_bu.cell(row=row, column=1).alignment = CENTER
            ws_bu.cell(row=row, column=2, value=t.get("tender_name", "")).font = BODY_FONT
            ws_bu.cell(row=row, column=3, value=t.get("org_name", "")).font = BODY_FONT
            ws_bu.cell(row=row, column=4, value=_format_budget(budget)).font = BODY_FONT
            ws_bu.cell(row=row, column=5, value=t.get("fit_score", 0)).font = BODY_FONT
            ws_bu.cell(row=row, column=5).alignment = CENTER
            ws_bu.cell(row=row, column=6, value=t.get("roi_score", 0)).font = BODY_FONT
            ws_bu.cell(row=row, column=6).alignment = CENTER
            ws_bu.cell(row=row, column=7, value="、".join(tag.get("tech_tags", []))).font = BODY_FONT
            ws_bu.cell(row=row, column=8, value=contact_text).font = BODY_FONT

            link_cell = ws_bu.cell(row=row, column=9, value="查看" if url else "")
            if url:
                link_cell.hyperlink = url
                link_cell.font = LINK_FONT
                link_cell.alignment = CENTER

            ws_bu.row_dimensions[row].height = 35

    # Save
    cfg = load_config(config_path)
    output_dir = Path(cfg["output"]["reports_dir"])
    output_dir.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"商機報告_{timestamp}.xlsx"
    wb.save(output_path)
    return str(output_path)


if __name__ == "__main__":
    data, source = load_latest_matched()
    path = create_excel_report(data, source)
    print(f"Excel 報告已產出: {path}")
